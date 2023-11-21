import os
from pathlib import Path
from typing import Union
import oyaml as yaml
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm
import openpyxl
from utils.manipulation_tools import merge_dict, map_dict_level, flatten_dict


###################################

class CSV_Generator(openpyxl.Workbook):
    """
    Generate a CSV file given a validation file and a timer file
    """

    def __init__(self, paths: Union[list, str, Path]):
        super().__init__()
        if isinstance(paths, str) or isinstance(paths, Path):
            files = [f for f in os.listdir(paths) if os.path.isfile(f)]
            if 'dataset.yaml' in files or 'Execution_time.yaml' in files or 'Validation.yaml' in files:
                self.paths = [paths]
            else:
                self.paths = [paths + '/' + f for f in os.listdir(paths) if os.path.isdir(paths + '/' + f)]
        else:
            self.paths = paths
        self.dataset = None
        self.validation = None
        self.timer = None
        self.create_dict_from_path()
        if self.dataset is not None:
            self.create_dataset_sheet()
        if self.validation is not None:
            self.create_validation_sheet()
        if self.timer is not None:
            self.create_timer_sheet()

    def create_dict_from_path(self):
        try:
            dataset = []
            for i in range(len(self.paths)):
                with open(self.paths[0] + '/dataset.yaml', 'r') as file:
                    dataset.append(yaml.safe_load(file))
            assert all(d == dataset[0] for d in dataset), f'The chosen experiments doesn"t feature the same dataset'
            self.dataset = dataset[0]
        except FileNotFoundError:
            print(f'There is no dataset file')

        try:
            validation = []
            for p in self.paths:
                with open(p + '/Validation.yaml', 'r') as file:
                    validation.append(yaml.safe_load(file)['2. results'])
            if len(validation) > 1:
                self.validation = merge_dict(*validation)
            else:
                self.validation = validation
        except FileNotFoundError:
            print(f'There is no Validation file')

        try:
            timer = []
            for p in self.paths:
                with open(p + '/Execution_time.yaml', 'r') as file:
                    timer.append({'Time per module': yaml.safe_load(file)['3. Time per module']})
            if len(timer) > 1:
                self.timer = merge_dict(*timer)
            else:
                self.timer = timer
        except FileNotFoundError:
            print(f'There is no Validation file')

    def create_dataset_sheet(self):
        sheet = self.create_sheet("Dataset")
        sheet.append([k for k in self.dataset['Files'].keys()])
        for nb in tqdm(range(self.dataset['Number of sample']),
                       total=self.dataset['Number of sample'],
                       desc=f"Dataset sheet creation"):
            sheet.append([self.dataset['Files'][k][nb] for k in self.dataset['Files'].keys()])
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

    def create_validation_sheet(self):
        for exp in self.validation.keys():
            # sheet = self.create_sheet(f"Validation_{exp}")
            for err in self.validation[exp].keys():
                sheet = self.create_sheet(f"{err}_{exp}")
                row = self.merge_according_dict(sheet, self.validation[exp][err])
                value_lists = flatten_dict(self.validation[exp][err])
                value_lists = [[col[i] for col in value_lists] for i in range(len(value_lists[0]))]
                for col, v in enumerate(value_lists):
                    col = get_column_letter(col+1)
                    sheet.sheet_view.selection[0].activeCell = f'{col}{row + 1}'
                    sheet.sheet_view.selection[0].sqref = f'{col}{row + 1}'
                    sheet.append(v)

    def create_timer_sheet(self):
        for exp in self.timer.keys():
            sheet = self.create_sheet(f"{exp}")
            row = self.merge_according_dict(sheet, self.timer[exp])
            value_lists = flatten_dict(self.timer[exp])
            # value_lists = [[col[i] for col in value_lists] for i in range(len(value_lists[0]))]
            sheet.sheet_view.selection[0].activeCell = f'{0}{row + 1}'
            sheet.sheet_view.selection[0].sqref = f'{0}{row + 1}'
            sheet.append(value_lists)

    def merge_according_dict(self, sheet, d: dict):
        merge_list, name_list = map_dict_level(d, level=0, map_of_dict=[], map_of_keys=[])
        for i in reversed(range(len(merge_list)-1)):
            temp = []
            map_dict_temp = merge_list.copy()
            for j in merge_list[i]:
                temp.append(sum(map_dict_temp[i+1][:j]))
                map_dict_temp[i + 1] = map_dict_temp[i + 1][j:]
            merge_list[i] = temp
        merge_list.append([1 for i in range(sum(merge_list[-1]))])
        for row in range(len(merge_list)):
            offset = 0
            for idx, range_merge in enumerate(merge_list[row]):
                sheet.merge_cells(start_row=row + 1, start_column=offset + 1,
                                  end_row=row + 1, end_column=offset + range_merge)
                c = sheet.cell(row=row + 1, column=offset + 1)
                c.value = name_list[row][idx % len(name_list[row])]
                c.alignment = Alignment(horizontal="center", vertical="center")
                offset += range_merge
        return row + 1


if __name__ == '__main__':
    exp = 'Resolution_comparison'
    gen = CSV_Generator(f'/home/godeta/PycharmProjects/Disparity_Pipeline/results/{exp}')
    gen.save(f"/home/godeta/PycharmProjects/Disparity_Pipeline/results/{exp}/results.xlsx")
